#!/usr/bin/env python3
"""
Excel Converter - Excel 檔案處理和轉換工具

功能：
- Excel 讀取與寫入
- 工作表操作
- 轉換為 CSV/JSON
- 批次處理多個工作簿
- 資料驗證
- 格式保留
"""

import argparse
import sys
import json
from pathlib import Path
from typing import List, Dict, Any, Union
import pandas as pd
import openpyxl
from openpyxl import load_workbook

def load_excel(file_path: str, sheet_name: Union[str, int] = 0) -> pd.DataFrame:
    """載入 Excel 檔案"""
    try:
        return pd.read_excel(file_path, sheet_name=sheet_name)
    except FileNotFoundError:
        print(f"❌ 檔案不存在: {file_path}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ 讀取錯誤: {e}")
        sys.exit(1)

def get_sheet_names(file_path: str) -> List[str]:
    """取得所有工作表名稱"""
    try:
        xl_file = pd.ExcelFile(file_path)
        return xl_file.sheet_names
    except Exception as e:
        print(f"❌ 讀取工作表失敗: {e}")
        return []

def excel_info(file_path: str):
    """顯示 Excel 檔案資訊"""
    print(f"📊 檔案資訊: {file_path}\n")

    sheets = get_sheet_names(file_path)
    print(f"📋 工作表數量: {len(sheets)}")

    for i, sheet in enumerate(sheets, 1):
        print(f"\n{i}. {sheet}")
        try:
            df = load_excel(file_path, sheet)
            print(f"   資料筆數: {len(df)}")
            print(f"   欄位數: {len(df.columns)}")
            print(f"   欄位名稱: {', '.join(df.columns[:5])}", end='')
            if len(df.columns) > 5:
                print(f" ... ({len(df.columns) - 5} 個欄位未顯示)")
            else:
                print()
        except Exception as e:
            print(f"   ❌ 讀取失敗: {e}")

def excel_to_csv(file_path: str, output_file: str, sheet_name: Union[str, int] = 0):
    """將 Excel 轉換為 CSV"""
    df = load_excel(file_path, sheet_name)
    df.to_csv(output_file, index=False, encoding='utf-8')
    print(f"✅ 已轉換為 CSV: {output_file}")
    print(f"   資料筆數: {len(df)}, 欄位數: {len(df.columns)}")

def excel_to_json(file_path: str, output_file: str, sheet_name: Union[str, int] = 0, orient: str = 'records'):
    """將 Excel 轉換為 JSON"""
    df = load_excel(file_path, sheet_name)
    df.to_json(output_file, orient=orient, force_ascii=False, indent=2)
    print(f"✅ 已轉換為 JSON: {output_file}")
    print(f"   資料筆數: {len(df)}, 欄位數: {len(df.columns)}")

def merge_sheets(file_path: str, output_file: str, output_format: str = 'csv'):
    """合併所有工作表"""
    sheets = get_sheet_names(file_path)

    if not sheets:
        print("❌ 沒有工作表可合併")
        return

    print(f"🔄 合併 {len(sheets)} 個工作表...")

    all_data = []
    for sheet in sheets:
        try:
            df = load_excel(file_path, sheet)
            df['來源工作表'] = sheet
            all_data.append(df)
            print(f"  ✅ {sheet}: {len(df)} 筆")
        except Exception as e:
            print(f"  ❌ {sheet}: 讀取失敗 - {e}")

    if not all_data:
        print("❌ 沒有資料可合併")
        return

    merged_df = pd.concat(all_data, ignore_index=True)
    print(f"\n✅ 合併完成，共 {len(merged_df)} 筆資料")

    if output_format == 'csv':
        merged_df.to_csv(output_file, index=False, encoding='utf-8')
    elif output_format == 'json':
        merged_df.to_json(output_file, orient='records', force_ascii=False, indent=2)
    elif output_format == 'excel':
        merged_df.to_excel(output_file, index=False)

    print(f"✅ 已儲存: {output_file}")

def split_sheets(file_path: str, output_dir: str, output_format: str = 'csv'):
    """將每個工作表分別儲存"""
    sheets = get_sheet_names(file_path)

    if not sheets:
        print("❌ 沒有工作表")
        return

    # 創建輸出目錄
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    print(f"📂 將 {len(sheets)} 個工作表分別儲存到: {output_dir}")

    for sheet in sheets:
        try:
            df = load_excel(file_path, sheet)

            # 清理檔名（移除特殊字元）
            safe_name = "".join(c for c in sheet if c.isalnum() or c in (' ', '-', '_')).strip()

            if output_format == 'csv':
                output_file = output_path / f"{safe_name}.csv"
                df.to_csv(output_file, index=False, encoding='utf-8')
            elif output_format == 'json':
                output_file = output_path / f"{safe_name}.json"
                df.to_json(output_file, orient='records', force_ascii=False, indent=2)
            elif output_format == 'excel':
                output_file = output_path / f"{safe_name}.xlsx"
                df.to_excel(output_file, index=False)

            print(f"  ✅ {sheet} -> {output_file.name} ({len(df)} 筆)")
        except Exception as e:
            print(f"  ❌ {sheet}: 失敗 - {e}")

def batch_convert(pattern: str, output_dir: str, output_format: str = 'csv'):
    """批次轉換多個 Excel 檔案"""
    from glob import glob

    files = glob(pattern)

    if not files:
        print(f"❌ 找不到符合的檔案: {pattern}")
        return

    # 創建輸出目錄
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    print(f"🔄 批次轉換 {len(files)} 個檔案...")

    for file_path in files:
        try:
            file_name = Path(file_path).stem

            if output_format == 'csv':
                output_file = output_path / f"{file_name}.csv"
                excel_to_csv(file_path, str(output_file))
            elif output_format == 'json':
                output_file = output_path / f"{file_name}.json"
                excel_to_json(file_path, str(output_file))

            print(f"  ✅ {file_name}")
        except Exception as e:
            print(f"  ❌ {file_name}: {e}")

def add_sheet(file_path: str, sheet_name: str, data_file: str):
    """新增工作表"""
    try:
        # 載入資料
        if data_file.endswith('.csv'):
            df = pd.read_csv(data_file)
        elif data_file.endswith('.json'):
            df = pd.read_json(data_file)
        else:
            print(f"❌ 不支援的檔案格式: {data_file}")
            return

        # 載入或創建工作簿
        try:
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        except FileNotFoundError:
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

        print(f"✅ 已新增工作表: {sheet_name}")
        print(f"   資料筆數: {len(df)}, 欄位數: {len(df.columns)}")
    except Exception as e:
        print(f"❌ 新增工作表失敗: {e}")

def remove_sheet(file_path: str, sheet_name: str):
    """移除工作表"""
    try:
        wb = load_workbook(file_path)

        if sheet_name not in wb.sheetnames:
            print(f"❌ 工作表不存在: {sheet_name}")
            return

        wb.remove(wb[sheet_name])
        wb.save(file_path)
        print(f"✅ 已移除工作表: {sheet_name}")
    except Exception as e:
        print(f"❌ 移除工作表失敗: {e}")

def main():
    parser = argparse.ArgumentParser(
        description='Excel Converter - Excel 檔案處理和轉換工具',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    parser.add_argument('input', nargs='?', help='輸入 Excel 檔案')
    parser.add_argument('--info', action='store_true', help='顯示 Excel 檔案資訊')
    parser.add_argument('--sheet', type=str, default='0', help='指定工作表（名稱或索引）')
    parser.add_argument('--to-csv', type=str, metavar='OUTPUT', help='轉換為 CSV')
    parser.add_argument('--to-json', type=str, metavar='OUTPUT', help='轉換為 JSON')
    parser.add_argument('--merge-sheets', action='store_true', help='合併所有工作表')
    parser.add_argument('--split-sheets', type=str, metavar='DIR', help='將工作表分別儲存')
    parser.add_argument('--batch', type=str, metavar='PATTERN', help='批次轉換（例如: *.xlsx）')
    parser.add_argument('--add-sheet', type=str, metavar='NAME', help='新增工作表')
    parser.add_argument('--data-file', type=str, help='資料檔案（用於新增工作表）')
    parser.add_argument('--remove-sheet', type=str, metavar='NAME', help='移除工作表')
    parser.add_argument('--format', choices=['csv', 'json', 'excel'], default='csv', help='輸出格式')
    parser.add_argument('-o', '--output', type=str, help='輸出檔案路徑')

    args = parser.parse_args()

    # 批次轉換
    if args.batch:
        output_dir = args.output or 'converted'
        batch_convert(args.batch, output_dir, args.format)
        return

    # 檢查輸入檔案
    if not args.input:
        parser.print_help()
        return

    # 新增工作表
    if args.add_sheet:
        if not args.data_file:
            print("❌ 請指定資料檔案 --data-file")
            return
        add_sheet(args.input, args.add_sheet, args.data_file)
        return

    # 移除工作表
    if args.remove_sheet:
        remove_sheet(args.input, args.remove_sheet)
        return

    # 顯示資訊
    if args.info:
        excel_info(args.input)
        return

    # 合併工作表
    if args.merge_sheets:
        output = args.output or f"{Path(args.input).stem}_merged.{args.format}"
        merge_sheets(args.input, output, args.format)
        return

    # 分割工作表
    if args.split_sheets:
        split_sheets(args.input, args.split_sheets, args.format)
        return

    # 處理工作表參數
    try:
        sheet = int(args.sheet)
    except ValueError:
        sheet = args.sheet

    # 轉換為 CSV
    if args.to_csv:
        excel_to_csv(args.input, args.to_csv, sheet)
        return

    # 轉換為 JSON
    if args.to_json:
        excel_to_json(args.input, args.to_json, sheet)
        return

    # 預設顯示資訊
    excel_info(args.input)

if __name__ == '__main__':
    main()
