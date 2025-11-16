"""
BOM Manager
智能 PCB 物料清單管理工具
"""

import csv
from pathlib import Path
from typing import List, Dict, Optional
from collections import defaultdict
from datetime import datetime


class BOMItem:
    """BOM 項目"""

    def __init__(
        self,
        references: List[str],
        value: str,
        footprint: str,
        mpn: str = "",
        manufacturer: str = "",
        description: str = ""
    ):
        self.references = sorted(references)
        self.value = value
        self.footprint = footprint
        self.mpn = mpn
        self.manufacturer = manufacturer
        self.description = description
        self.quantity = len(references)
        self.pricing = {}
        self.stock = None
        self.supplier = ""
        self.sku = ""

    @property
    def reference_str(self) -> str:
        """獲取引用字串"""
        return ",".join(self.references)

    def __repr__(self):
        return f"BOMItem({self.quantity}x {self.value} {self.footprint})"


class BOMManager:
    """BOM 管理器"""

    def __init__(self):
        self.items: List[BOMItem] = []
        self.metadata = {
            'project_name': '',
            'version': '',
            'date': datetime.now().isoformat(),
            'tool': 'BOM Manager v0.1.0'
        }

    def extract_from_kicad(self, pcb_file: str) -> None:
        """
        從 KiCAD PCB 檔案提取 BOM

        Args:
            pcb_file: KiCAD PCB 檔案路徑
        """
        try:
            import pcbnew
        except ImportError:
            raise ImportError("需要 pcbnew 模組,請在 KiCAD 環境中執行")

        print(f"📋 從 KiCAD 提取 BOM: {pcb_file}")

        board = pcbnew.LoadBoard(pcb_file)
        self.metadata['project_name'] = Path(pcb_file).stem

        # 收集元件資訊
        components = defaultdict(list)

        for fp in board.GetFootprints():
            ref = fp.GetReference()
            value = fp.GetValue()
            footprint = str(fp.GetFPID().GetLibItemName())

            # 獲取額外屬性
            mpn = ""
            manufacturer = ""

            # 嘗試從屬性中獲取 MPN 和製造商
            for field in fp.GetFields():
                field_name = field.GetName().lower()
                field_value = field.GetText()

                if 'mpn' in field_name or 'part' in field_name:
                    mpn = field_value
                elif 'manufacturer' in field_name or 'mfr' in field_name:
                    manufacturer = field_value

            # 組合鍵: (值, 封裝, MPN)
            key = (value, footprint, mpn)
            components[key].append(ref)

        # 建立 BOM 項目
        self.items = []
        for (value, footprint, mpn), refs in components.items():
            item = BOMItem(
                references=refs,
                value=value,
                footprint=footprint,
                mpn=mpn
            )
            self.items.append(item)

        # 按引用排序
        self.items.sort(key=lambda x: x.references[0])

        print(f"✅ 提取完成: {self.total_components} 個元件, {self.unique_components} 種類")

    def extract_from_csv(self, csv_file: str) -> None:
        """
        從 CSV 檔案載入 BOM

        Args:
            csv_file: CSV 檔案路徑
        """
        print(f"📋 從 CSV 載入 BOM: {csv_file}")

        self.items = []

        with open(csv_file, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)

            for row in reader:
                refs = row.get('References', '').split(',')
                refs = [r.strip() for r in refs if r.strip()]

                item = BOMItem(
                    references=refs,
                    value=row.get('Value', ''),
                    footprint=row.get('Footprint', ''),
                    mpn=row.get('MPN', ''),
                    manufacturer=row.get('Manufacturer', ''),
                    description=row.get('Description', '')
                )

                # 載入價格資訊(如果有)
                if 'Unit Price' in row:
                    try:
                        item.pricing[1] = float(row['Unit Price'].replace('$', ''))
                    except:
                        pass

                self.items.append(item)

        print(f"✅ 載入完成: {self.total_components} 個元件, {self.unique_components} 種類")

    def optimize(
        self,
        merge_duplicates: bool = True,
        standardize_names: bool = True
    ) -> None:
        """
        優化 BOM

        Args:
            merge_duplicates: 合併重複元件
            standardize_names: 標準化元件名稱
        """
        print("🔧 優化 BOM...")

        original_count = len(self.items)

        if merge_duplicates:
            self._merge_duplicates()

        if standardize_names:
            self._standardize_names()

        optimized_count = len(self.items)
        print(f"✅ 優化完成: {original_count} → {optimized_count} 項目")

    def _merge_duplicates(self) -> None:
        """合併重複的元件"""
        merged = defaultdict(list)

        for item in self.items:
            key = (item.value, item.footprint, item.mpn)
            merged[key].append(item)

        self.items = []
        for items in merged.values():
            if len(items) == 1:
                self.items.append(items[0])
            else:
                # 合併多個項目
                all_refs = []
                for item in items:
                    all_refs.extend(item.references)

                merged_item = BOMItem(
                    references=all_refs,
                    value=items[0].value,
                    footprint=items[0].footprint,
                    mpn=items[0].mpn,
                    manufacturer=items[0].manufacturer,
                    description=items[0].description
                )
                self.items.append(merged_item)

    def _standardize_names(self) -> None:
        """標準化元件名稱"""
        # 標準化電阻、電容等的值
        for item in self.items:
            # 電阻值標準化 (例如: 10K → 10kΩ)
            if any(item.references[0].startswith(prefix) for prefix in ['R', 'r']):
                item.value = self._standardize_resistance(item.value)

            # 電容值標準化
            elif any(item.references[0].startswith(prefix) for prefix in ['C', 'c']):
                item.value = self._standardize_capacitance(item.value)

    def _standardize_resistance(self, value: str) -> str:
        """標準化電阻值"""
        value = value.upper()
        if 'K' in value:
            return value.replace('K', 'kΩ')
        elif 'M' in value:
            return value.replace('M', 'MΩ')
        else:
            return value + 'Ω' if not value.endswith('Ω') else value

    def _standardize_capacitance(self, value: str) -> str:
        """標準化電容值"""
        value = value.lower()
        if 'u' in value or 'µ' in value:
            return value.replace('u', 'µF').replace('µ', 'µF')
        elif 'n' in value:
            return value.replace('n', 'nF')
        elif 'p' in value:
            return value.replace('p', 'pF')
        else:
            return value

    def export_csv(
        self,
        output_file: str,
        include_pricing: bool = False
    ) -> None:
        """
        輸出為 CSV 格式

        Args:
            output_file: 輸出檔案路徑
            include_pricing: 是否包含價格資訊
        """
        print(f"💾 輸出 CSV: {output_file}")

        headers = [
            'Item', 'Quantity', 'References', 'Value',
            'Footprint', 'MPN', 'Manufacturer', 'Description'
        ]

        if include_pricing:
            headers.extend(['Supplier', 'SKU', 'Unit Price', 'Total Price'])

        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(headers)

            for idx, item in enumerate(self.items, 1):
                row = [
                    idx,
                    item.quantity,
                    item.reference_str,
                    item.value,
                    item.footprint,
                    item.mpn,
                    item.manufacturer,
                    item.description
                ]

                if include_pricing:
                    unit_price = item.pricing.get(1, 0.0)
                    total_price = unit_price * item.quantity
                    row.extend([
                        item.supplier,
                        item.sku,
                        f"${unit_price:.3f}",
                        f"${total_price:.2f}"
                    ])

                writer.writerow(row)

        print(f"✅ CSV 輸出完成")

    def export_excel(
        self,
        output_file: str,
        include_pricing: bool = False,
        include_stock: bool = False
    ) -> None:
        """
        輸出為 Excel 格式

        Args:
            output_file: 輸出檔案路徑
            include_pricing: 是否包含價格資訊
            include_stock: 是否包含庫存資訊
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            print("⚠️  需要 openpyxl 套件: pip install openpyxl")
            return

        print(f"💾 輸出 Excel: {output_file}")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "BOM"

        # 標題行
        headers = [
            'Item', 'Quantity', 'References', 'Value',
            'Footprint', 'MPN', 'Manufacturer', 'Description'
        ]

        if include_pricing:
            headers.extend(['Supplier', 'SKU', 'Unit Price', 'Total Price'])

        if include_stock:
            headers.append('Stock')

        ws.append(headers)

        # 設定標題樣式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # 資料行
        for idx, item in enumerate(self.items, 1):
            row = [
                idx,
                item.quantity,
                item.reference_str,
                item.value,
                item.footprint,
                item.mpn,
                item.manufacturer,
                item.description
            ]

            if include_pricing:
                unit_price = item.pricing.get(1, 0.0)
                total_price = unit_price * item.quantity
                row.extend([
                    item.supplier,
                    item.sku,
                    unit_price,
                    total_price
                ])

            if include_stock:
                row.append(item.stock or 0)

            ws.append(row)

        # 調整欄寬
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 20

        wb.save(output_file)
        print(f"✅ Excel 輸出完成")

    @property
    def total_components(self) -> int:
        """總元件數"""
        return sum(item.quantity for item in self.items)

    @property
    def unique_components(self) -> int:
        """唯一元件種類數"""
        return len(self.items)

    @property
    def estimated_cost(self) -> float:
        """估計成本"""
        total = 0.0
        for item in self.items:
            unit_price = item.pricing.get(1, 0.0)
            total += unit_price * item.quantity
        return total


if __name__ == "__main__":
    # 簡單測試
    print("BOM Manager")
    print("使用範例:")
    print("""
    bom = BOMManager()
    bom.extract_from_kicad('board.kicad_pcb')
    bom.optimize()
    bom.export_csv('bom_output.csv')
    bom.export_excel('bom_output.xlsx')
    """)
